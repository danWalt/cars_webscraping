import gspread
from oauth2client.service_account import ServiceAccountCredentials
import requests
from bs4 import BeautifulSoup
import openpyxl

# Google APIs scope
scope = ["https://spreadsheets.google.com/feeds",
         "https://www.googleapis.com/auth/spreadsheets",
         "https://www.googleapis.com/auth/drive.file",
         "https://www.googleapis.com/auth/drive"]

# Creds to use api
creds = ServiceAccountCredentials.from_json_keyfile_name('creds.json', scope)

# load json and scope
client = gspread.authorize(creds)

# load active sheet
sheets = client.open('cars').sheet1

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36'}

wb = openpyxl.load_workbook('cars.xlsx')
s = wb.active

cars = []
row = 2
page_num = 361


while page_num < 838:
    print('processing page number ' + str(page_num))
    # yad 2 cars page
    URL = 'https://www.yad2.co.il/vehicles/private-cars?page=%s' % page_num
    # Load yad2 cars page
    page = requests.get(URL, headers=headers)
    page.raise_for_status()  # make sure page loaded without problems

    # load page content with BS
    soup = BeautifulSoup(page.content, 'html.parser')

    # list of car descriptions
    descriptions_list = soup.find_all(class_='subtitle')

    # organize ad details and add a row for each car to the google sheet
    for i in range(0, 40):
        try:
            car_type = soup.find(
                id='feed_item_%s_title' % i).get_text().strip()  # car type
            description = descriptions_list[
                i].get_text()  # car description from ad
            year = soup.find(
                id='data_year_%s' % i).get_text()  # car year manufactured
            hand = soup.find(
                id='data_hand_%s' % i).get_text()  # Car number of previous owners
            engine = soup.find(
                id='data_engine_size_%s' % i).get_text()  # car engine capaity
            price = soup.find(id='feed_item_%s_price' % i).get_text().strip()
            price = price[:len(price) - 2]  # car price in NIS
            cars.append([car_type, description, year, hand, engine, price])
        except AttributeError:
            for i in range(len(cars)):
                r = s.max_row + 1
                print("attribute error, adding to sheet car num " + str(row))
                for j in range(0, len(cars[i])):
                    s.cell(r, j + 1, cars[i][j])
                row += 1
            print('page num ' + str(page_num))
            page_num += 9999
            break


    page_num += 1
#
wb.save('cars.xlsx')
