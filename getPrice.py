import requests
from bs4 import BeautifulSoup
import openpyxl
import pprint

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36'}

URL = 'http://pricelist.yad2.co.il/'

# Load yad2 cars prices page
page = requests.get(URL, headers=headers)
page.raise_for_status()  # make sure page loaded without problems

# load page content with BS
soup = BeautifulSoup(page.content, 'html.parser')

# create car types dictionary text -> site code
car_types = soup.select('select[name=CarType] > option')
car_dic = {}
for i in car_types:
    # print("mmmm")
    # print(i)
    car_dic[i.text] = i['value']
    # print(car_dic[i.text])
# create manufacturer dictionary text -> site code
manufacturer = soup.select('select[name=Manufactur] > option')
manufacturer_dic = {}
for i in manufacturer:
    manufacturer_dic[i.text] = i['value']
pprint.pprint(manufacturer)
pprint.pprint(manufacturer_dic)


# create model dictionary text -> site code
model = soup.select('select[name=CarModel] > option')
mod_dic = {}
for i in model:
    mod_dic[i.text] = i['value']

# create year dictionary text -> site code
year = soup.select('select[name=Year] > option')
year_dic = {}
for i in year:
    year_dic[i.text] = i['value']

wb = openpyxl.load_workbook('cars.xlsx')
s = wb.active

pprint.pprint(manufacturer_dic)

for row in range(2, s.max_row):
    manu = manufacturer_dic[s.cell(row, 1).value.strip()]
    mod = mod_dic[s.cell(row, 2).value]
    y = year_dic[s.cell(row, 4).value]

    searchURL = 'http://pricelist.yad2.co.il/search.php?CarType=1&Manufactur' \
                '=%s&CarModel=%s&Year=%s' % (manu, mod, y)

    page2 = requests.get(searchURL, headers=headers)

    page2.raise_for_status()  # make sure page loaded without problems

    # load page content with BS
    soup = BeautifulSoup(page2.content, 'html.parser')

    p = soup.find_all("td", class_="grey_bold")
    price = p[1].get_text().strip()[:6]
    s.cell(row, 8).value = price
    s.cell(row, 9).value = 1 - (int(s.cell(row, 8).value) / s.cell(row,
                                                                   7).value)
